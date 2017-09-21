#!/usr/bin/env python
# encoding: utf-8

""" Script para analizar los logs de la aplicación Terminkalender en su versión en papel.
    Visita la documentación en GitHub para más información.
    Autor: Luis Gonzaga Rozo Bueno
"""

from openpyxl import load_workbook

import os
import re
import enchant


def aplicar_preprocesado(msg):
    laughing = re.findall(r"jaj.*|jej.*|jij.*|joj.*|juj.*|hah.*", msg, re.IGNORECASE)
    if laughing:
        for replacement in laughing:
            msg = msg.replace(replacement, "jajaja")

    greetings = re.findall('hallo*', msg, re.IGNORECASE)
    if greetings:
        for replacement in greetings:
            msg = msg.replace(replacement, "hallo")

    ja = re.findall('jaa+', msg, re.IGNORECASE)
    if ja:
        for replacement in ja:
            msg = msg.replace(replacement, "ja")

    anxious = re.findall(r'\?+!+', msg, re.IGNORECASE)
    if anxious:
        for replacement in anxious:
            msg = msg.replace(replacement, "?")

    sharp = re.findall(r'#+', msg, re.IGNORECASE)
    if sharp:
        for replacement in sharp:
            msg = msg.replace(replacement, "#")

    interrogation = re.findall(r'\?+', msg, re.IGNORECASE)
    if interrogation:
        for replacement in interrogation:
            msg = msg.replace(replacement, " ?")

    asterisk = re.findall('\*+', msg, re.IGNORECASE)
    if asterisk:
        for replacement in asterisk:
            msg = msg.replace(replacement, " * ")

    exclamation = re.findall('\!+', msg, re.IGNORECASE)
    if exclamation:
        for replacement in exclamation:
            msg = msg.replace(replacement, " !")

    ellipsis = re.findall('\.\.+', msg, re.IGNORECASE)
    if ellipsis:
        for replacement in ellipsis:
            msg = msg.replace(replacement, " ... ")

    comma = re.findall(',\s?', msg, re.IGNORECASE)
    if comma:
        for replacement in comma:
            msg = msg.replace(replacement, " , ")

    a_vowel = re.findall('aa+', msg, re.IGNORECASE)
    if a_vowel:
        for replacement in a_vowel:
            msg = msg.replace(replacement, "a")

    e_vowel = re.findall('ee+', msg, re.IGNORECASE)
    if e_vowel:
        for replacement in e_vowel:
            msg = msg.replace(replacement, "e")

    i_vowel = re.findall('ii+', msg, re.IGNORECASE)
    if i_vowel:
        for replacement in i_vowel:
            msg = msg.replace(replacement, "i")

    o_vowel = re.findall('oo+', msg, re.IGNORECASE)
    cool = re.findall('cool', msg, re.IGNORECASE)
    if o_vowel and not cool:
        for replacement in o_vowel:
            msg = msg.replace(replacement, "o")

    u_vowel = re.findall('uu+', msg, re.IGNORECASE)
    if u_vowel:
        for replacement in u_vowel:
            msg = msg.replace(replacement, "u")

    marks = re.findall('"', msg, re.IGNORECASE)
    if marks:
        for replacement in marks:
            msg = msg.replace(replacement, "")

    l_paren = re.findall('\(', msg, re.IGNORECASE)
    if l_paren:
        for replacement in l_paren:
            msg = msg.replace(replacement, " ( ")

    r_paren = re.findall('\)', msg, re.IGNORECASE)
    if r_paren:
        for replacement in r_paren:
            msg = msg.replace(replacement, " ) ")

    point = re.findall('\.', msg, re.IGNORECASE)
    shocked = re.findall('o.o', msg, re.IGNORECASE)
    if point and not ellipsis and not shocked:
        for replacement in point:
            msg = msg.replace(replacement, " .")

    return msg


class Mensaje:
    def __init__(self, mensaje, alumno, pareja, cambio_turno):
        self.mensaje = mensaje
        self.alumno = alumno
        self.pareja = pareja
        self.cambio_turno = cambio_turno

    def __str__(self):
        print "--------------------------"
        print "Mensaje: " + self.mensaje
        print "Alumno: " + self.alumno
        print "Pareja: " + self.pareja
        print "--------------------------"


alumnos = []
parejas = []


def obtener_mensajes(maximo_filas):
    mensajes_hoja = []
    i = 0
    alumno_izqda = ""
    alumno_dcha = ""
    pareja = ""

    while i < maximo_filas:
        if data[i][0].font.b and data[i][1].font.b and data[i][0].value is not None and data[i][1].value is not None:
            alumno_izqda = data[i][0].value.encode("utf-8")
            alumno_dcha = data[i][1].value.encode("utf-8")
            pareja = alumno_izqda + "-" + alumno_dcha

            if alumno_izqda not in alumnos:
                alumnos.append(alumno_izqda)
            if alumno_dcha not in alumnos:
                alumnos.append(alumno_dcha)
            if pareja not in parejas:
                parejas.append(pareja)

        elif not data[i][0].font.b and type(data[i][0]) != 'NoneType':
            if data[i][0].value is not None:
                cambio_turno = False
                if data[i-1][0].font.b or data[i-1][0].value is None:
                    cambio_turno = True
                msg_prep = aplicar_preprocesado(data[i][0].value.encode("utf-8"))
                mensaje_final = Mensaje(msg_prep, alumno_izqda, pareja, cambio_turno)
                mensajes_hoja.append(mensaje_final)

        if not data[i][1].font.b and type(data[i][1]) != 'NoneType':
            if data[i][1].value is not None:
                cambio_turno = False
                if data[i-1][1].font.b or data[i-1][1].value is None:
                    cambio_turno = True
                msg_prep = aplicar_preprocesado(data[i][1].value.encode("utf-8"))
                mensaje_final = Mensaje(msg_prep, alumno_dcha, pareja, cambio_turno)
                mensajes_hoja.append(mensaje_final)

        i += 1

    return mensajes_hoja

# Abrimos todas las hojas de cálculo en las que se van a buscar las estadísticas
ficheros = []

for fichero in os.listdir("."):
    if ".xlsx" in fichero and "#" not in fichero and "Agenda" not in fichero:
        ficheros.append(fichero)

Mensajes = []
for fichero in ficheros:
    prueba = load_workbook(fichero)

    for hoja in prueba:
        inicio = 'A1'
        fin = 'B' + str(hoja.max_row)
        intervalo = inicio + ':' + fin
        data = hoja[intervalo]

        # Mensajes
        maximo_filas = hoja.max_row
        msgs = obtener_mensajes(maximo_filas)
        for m in msgs:
            Mensajes.append(m)

# Para obtener las estadisticas referentes a palabras o expresiones reservadas, se hace uso de la siguientes estructuras

pronombres_interrogativos = ["mit wem", "um wie viel Uhr", "um wieviel Uhr", "wann", "was", "wer", "wie",
                             "wo", "wohin"]

adjetivos = ["frei", "gestresst", "hungrig", "müde", "happy", "glücklich"]

saludos_despedidas = ["auf Wiedersehen!", "bis Montag", "bis Dienstag",
                      "bis Mittwoch", "bis Donnerstag", "bis Freitag", "bis Samstag", "bis Sonntag",
                      "bis bald", "bis dann", "bis später",
                      "ciao", "hallihallo", "hallo", "hey", "hi", "mach's gut", "tschüs",
                      "tschüs, bis bald", "tschüs, bis dann", "tschüs, bis später", "tschüss",
                      "tschüss, bis bald",
                      "tschüss, bis dann", "tschüss, bis später",
                      "tschüs bis bald", "tschüs bis dann", "tschüs bis später", "tschüss bis bald",
                      "tschüss bis dann", "tschüss bis später"]

frases_hechas = ["ich bin gestresst", "ich bin happy", "ich bin relaxt", "ich bin hungirg",
                 "ich bin frustriert",
                 "ich bin deprimiert", "ich bin alleine", "ich bin müde", "ich bin fit",
                 "ich bin total fit",
                 "ich habe Zeit",
                 "ich habe keine Zeit",
                 "ich habe leider keine Zeit",
                 "mir geht's gut", "mir geht's schlecht", "mir geht's so lala", "mir geht's fatal",
                 "mir geht's super",
                 "mir geht's total super", "mir geht's phantastisch", "mir geht's fantastisch",
                 "mir geht's wunderbar",
                 "mir geht's genial", "mir geht's total gut", "mir geht's total schlecht",
                 "mir geht's spitze"]

expresiones_afirmacion = ["alles klar", "das ist gut", "genial", "gut", "ja, cool", "ja cool", "ja, klar",
                          "ja klar",
                          "ja, klasse", "ja klasse", "ja, natürlich", "ja natürlich", "ja, perfekt",
                          "ja perfekt",
                          "ja, phantastisch", "ja phantastisch", "ja, super Idee", "ja super Idee",
                          "ja, super", "ja super", "ja, wunderbar", "ja wunderbar", "kein Problem", "klasse",
                          "ok",
                          "okay", "super", "total super"]

expresiones_negacion_duda = ["nein", "ne", "nee", "neee", "keine Ahnung", "ich weiss nicht",
                             "ich habe keine Ahnung", "hm"]

lugares = ["Bibliothek", "Café", "Copyshop", "Copycenter", "Disco", "Disko", "Diskothek", "Einkaufszentrum",
           "Eiscafé", "Fitnessstudio", "Fitnesszentrum", "Fussballplatz", "Fußballplatz", "Fussballstadion",
           "Fußballstadion", "Fussballstadium", "Fußballstadium", "Kino", "Kneipe", "Konzert", "Pub",
           "Restaurant",
           "Schwimmbad", "Sportzentrum", "Strand", "Supermarkt", "Tennisplatz", "Theater", "Uni",
           "Universität",
           "Zentrum", "zu Hause", "zuhause"]

dias_de_la_semana = ["Mittwoch", "Donnerstag", "Freitag", "Samstag", "Sonntag", "Wochenende"]

adverbios_temporales = ["am Montagmorgen", "am Montagvormittag", "am Montagmittag",
                        "am Montagnachmittag", "am Montagabend", "am Dienstagmorgen", "am Dienstagvormittag",
                        "am Dienstagmittag", "am Dienstagnachmittag", "am Dienstagabend", "am Mittwochmorgen",
                        "am Mittwochvormittag", "am Mittwochmittag", "am Mittwochnachmittag",
                        "am Mittwochabend",
                        "am Donnerstagmorgen", "am Donnerstagvormittag", "am Donnerstagmittag",
                        "am Donnerstagnachmittag", "am Donnerstagabend", "am Freitagmorgen",
                        "am Freitagvormittag",
                        "am Freitagmittag", "am Freitagnachmittag", "am Freitagabend", "am Samstagmorgen",
                        "am Samstagvormittag", "am Samstagmittag", "am Samstagnachmittag", "am Samstagabend",
                        "am Sonntagmorgen", "am Sonntagvormittag", "am Sonntagmittag", "am Sonntagnachmittag",
                        "am Sonntagabend",
                        "Montagnacht", "Dienstagnacht", "Mittwochnacht", "Donnerstagnacht",
                        "Freitagnacht", "Samstagnacht", "Sonntagnacht",
                        'montagmorgens', 'montagvormittags', 'montagmittags', 'montagnachmittags',
                        'montagabends',
                        'montagnachts', 'dienstagmorgens', 'dienstagvormittags', 'dienstagmittags',
                        'dienstagnachmittags', 'dienstagabends', 'dienstagnachts', 'mittwochmorgens',
                        'mittwochvormittags', 'mittwochmittags', 'mittwochnachmittags', 'mittwochabends',
                        'mittwochnachts', 'donnerstagmorgens', 'donnerstagvormittags', 'donnerstagmittags',
                        'donnerstagnachmittags', 'donnerstagabends', 'donnerstagnachts', 'freitagmorgens',
                        'freitagvormittags', 'freitagmittags', 'freitagnachmittags', 'freitagabends',
                        'freitagnachts',
                        'samstagmorgens', 'samstagvormittags', 'samstagmittags', 'samstagnachmittags',
                        'samstagabends',
                        'samstagnachts', 'sonntagmorgens', 'sonntagvormittags', 'sonntagmittags',
                        'sonntagnachmittags',
                        'sonntagabends', 'sonntagnachts',
                        "am Vormittag", "am Mittag", "am Nachmittag", "am Abend", "in der Nacht", "am Montag",
                        "am Dienstag", "am Mittwoch", "am Donnerstag", "am Freitag", "am Samstag", "am Sonntag"]

horas = ["um 1 Uhr", "um 2 Uhr", "um 3 Uhr", "um 4 Uhr", "um 5 Uhr", "um 6 Uhr", "um 7 Uhr", "um 8 Uhr",
         "um 9 Uhr", "um 10 Uhr", "um 11 Uhr", "um 12 Uhr", "um 13 Uhr", "um 14 Uhr", "um 15 Uhr", "um 16 Uhr",
         "um 17 Uhr", "um 18 Uhr", "um 19 Uhr", "um 20 Uhr", "um 21 Uhr", "um 22 Uhr", "um 23 Uhr", "um 24 Uhr",
         'um 1', 'um 2', 'um 3', 'um 4', 'um 5', 'um 6', 'um 7', 'um 8', 'um 9', 'um 10', 'um 11', 'um 12',
         'um 13', 'um 14', 'um 15', 'um 16', 'um 17', 'um 18', 'um 19', 'um 20', 'um 21', 'um 22', 'um 23', 'um 24',
         "ein", "zwei", "drei", "vier", "fünf", "sechs", "sieben", "acht", "neun", "zehn", "elf", "zwölf",
         "dreizehn", "vierzehn", "fünfzehn", "sechzehn", "siebzehn", "achtzehn", "neunzehn", "zwanzig",
         "einundzwanzig",
         "zweiundzwanzig", "dreiundzwanzig", "vierundzwanzig",
         'um ein', 'um zwei', 'um drei', 'um vier', 'um fünf', 'um sechs', 'um sieben', 'um acht', 'um neun',
         'um zehn', 'um elf', 'um zwölf', 'um dreizehn', 'um vierzehn', 'um fünfzehn', 'um sechzehn',
         'um siebzehn', 'um achtzehn', 'um neunzehn', 'um zwanzig', 'um einundzwanzig', 'um zweiundzwanzig',
         'um dreiundzwanzig', 'um vierundzwanzig']

preposicion_mit = ["mit"]

preposiciones_con_pronombres_personales = ["mit mir", "mit dir", "mit uns"]

''' Nuevas estructuras '''

''' '''

# Si se añaden nuevos tipos de palabras o expresiones reservadas, añadirlo tambien en estas dos estructuras.

tipos_expresiones = ["Pronombre interrogativo", "Adjetivo", "Saludo o despedida", "Frase hecha",
                     "Expresion de afirmacion", "Expresion de negacion o duda", "Lugar", "Dia de la semana",
                     "Adverbio temporal", "Hora", "Preposicion mit", "Preposicion con pronombre personal"]

array_expresiones = [pronombres_interrogativos, adjetivos, saludos_despedidas, frases_hechas,
                     expresiones_afirmacion, expresiones_negacion_duda, lugares, dias_de_la_semana,
                     adverbios_temporales, horas, preposicion_mit, preposiciones_con_pronombres_personales]

# Para tener en cuenta las mayusculas o minusculas en las busquedas, se hara uso de esta lista de sustantivos:

sustantivos = ["Abend", "Ahnung", "Bibliothek", "Café", "Copycenter", "Copyshop", "Dienstag", "Dienstagnacht",
               "Disco",
               "Disko", "Diskothek", "Donnerstag", "Donnerstagnacht", "Einkaufszentrum", "Eiscafé",
               "Fitnessstudio",
               "Fitnesszentrum", "Freitag", "Freitagnacht", "Fussballplatz", "Fußballplatz", "Fussballstadion",
               "Fußballstadion", "Fussballstadium", "Fußballstadium", "Idee", "Kino", "Kneipe", "Konzert",
               "Mittag",
               "Mittwoch", "Mittwochnacht", "Montag", "Montagnacht", "Nachmittag", "Nacht", "Problem", "Pub",
               "Restaurant", "Samstag", "Samstagnacht", "Schwimmbad", "Sonntag", "Sonntagnacht", "Sportzentrum",
               "Strand", "Supermarkt", "Tennisplatz", "Theater", "Uhr", "Uni", "Universität", "Vormittag",
               "Wochenende",
               "Zeit", "Zentrum", "zu Hause", "zuhause"]


# Esta ampliación de diccionario contiene palabras que el módulo 'enchant' no
# considera correctas para el diccionario alemán, pero no son incorrectas al tratarse
# de nombres de lugares, actividades, etc., que se utilizan en los chats para concretar
# los planes.
ampliacion_diccionario = []
fichero_con_tareas = open(r'../tasks.xml', 'r').readlines()
for linea in fichero_con_tareas:
    if '<field>' in linea:
        palabras = re.sub(r'<field>|\(|\)|</field>', '', linea).split()
        for palabra in palabras:
            if palabra not in ampliacion_diccionario:
                ampliacion_diccionario.append(palabra)

for adv in adverbios_temporales:
    ampliacion_diccionario.append(adv)


def usaExpresionReservada(frase):
    # El array expresionesUtilizadas contendra tuplas de la forma [Tipo de expresion, expresion]
    expresionesUtilizadas = []
    # La variable 'i' ayuda a obtener el tipo de la expresion que se ha detectado
    i = 0
    for tipo_expresion in array_expresiones:
        for expresion in tipo_expresion:
            sustantivos_utilizados = []
            sustantivos_minusculas = []
            busqueda_refinada = False

            # Determinamos si hacer busqueda refinada, y guardamos los sustantivos (bien o mal) utilizados en la frase.
            for sustantivo in sustantivos:
                sustantivo_encontrado = re.findall(r'\b' + sustantivo + r'\b', frase, re.IGNORECASE)
                for sust in sustantivo_encontrado:
                    if sust.istitle():  # Sustantivo bien utilizado
                        sustantivos_utilizados.append(sust)
                        busqueda_refinada = True
                    else:
                        sustantivos_minusculas.append(sust)

            # Se pasa la frase a minusuculas, dejando solamente los sustantivos en mayuscula.
            if busqueda_refinada:
                frase = frase.lower()
                for sustantivo in sustantivos_utilizados:
                    sustitucion = re.findall(r'\b' + sustantivo + r'\b', frase, re.IGNORECASE)
                    frase = frase.replace(sustitucion[0], sustantivo)

                if expresion in frase:
                    expresionesUtilizadas.append([tipos_expresiones[i], expresion])

            # Se compara con la frase original o con la frase en minusculas, eliminando los sustantivos mal utilizados.
            else:
                frase_split = frase.split()
                for sust in sustantivos_minusculas:
                    frase_split.remove(sust)
                frase_completa = ' '.join(frase_split)

                expr = re.findall(r'\b'+expresion+r'\b', frase_completa, re.IGNORECASE)
                if expr:
                    expresionesUtilizadas.append([tipos_expresiones[i], expresion])
        i += 1

    return expresionesUtilizadas

stats_per_student = {}
stats_per_pair = {}

# Se utiliza un array para almacenar las estadisticas por estudiante y por pareja.
# Para ello, se le asigna dicho vector a cada entrada del diccionario, respondiendo a los siguientes indices:
# 0: Mensajes
# 1: Frases
# 2: Turnos (en unidades)
# 3: Palabras
# 4: Palabras en diccionario
# 5: Palabras fuera diccionario
# 6: Palabras distintas
# 7: Palabras en diccionario distintas
# 8: Palabras fuera diccionario distintas
# 9: Palabras reservadas total
# 10: Palabras reservadas - Categoría "Pronombres interrogativos"
# 11: Palabras reservadas - Categoría "Adjetivos"
# 12: Palabras reservadas - Categoría "Saludos y despedidas"
# 13: Palabras reservadas - Categoría "Frases hechas"
# 14: Palabras reservadas - Categoría "Expresiones de afirmación"
# 15: Palabras reservadas - Categoría "Expresiones de negación/duda"
# 16: Palabras reservadas - Categoría "Lugares"
# 17: Palabras reservadas - Categoría "Días de la semana"
# 18: Palabras reservadas - Categoría "Adverbios temporales"
# 19: Palabras reservadas - Categoría "Horas"
# 20: Palabras reservadas - Categoría "Preposición 'mit'"
# 21: Palabras reservadas - Categoría "'Mit' + pronombres personales"
# 22: Palabras reservadas distintas
# 23: Palabras reservadas distintas - Categoría "Pronombres interrogativos"
# 24: Palabras reservadas distintas - Categoría "Adjetivos"
# 25: Palabras reservadas distintas - Categoría "Saludos y despedidas"
# 26: Palabras reservadas distintas - Categoría "Frases hechas"
# 27: Palabras reservadas distintas - Categoría "Expresiones de afirmación"
# 28: Palabras reservadas distintas - Categoría "Expresiones de negación/duda"
# 29: Palabras reservadas distintas - Categoría "Lugares"
# 30: Palabras reservadas distintas - Categoría "Días de la semana"
# 31: Palabras reservadas distintas - Categoría "Adverbios temporales"
# 32: Palabras reservadas distintas - Categoría "Horas"
# 33: Palabras reservadas distintas - Categoría "Preposición 'mit'"
# 34: Palabras reservadas distintas - Categoría "'Mit' + pronombres personales"
# 35: Frases de una sola palabra
# 36: Frases interrogativas
# 37: Frases exclamativas
# 38: Estructuras gramaticales precisas total
# 39: Estructuras gramaticales precisas - Categoría "Preguntas con Complemento de Lugar"
# 40: Estructuras gramaticales precisas - Categoría "Preguntas con Complemento de Tiempo"
# 41: Estructuras gramaticales precisas - Categoría "Preguntas simples"
# 42: Estructuras gramaticales precisas - Categoría "Enunciados con Complemento Directo"
# 43: Estructuras gramaticales precisas distintas
# 44: Estructuras gramaticales precisas distintas - Categoría "Preguntas con Complemento de Lugar"
# 45: Estructuras gramaticales precisas distintas - Categoría "Preguntas con Complemento de Tiempo"
# 46: Estructuras gramaticales precisas distintas - Categoría "Preguntas simples"
# 47: Estructuras gramaticales precisas distintas - Categoría "Enunciados con Complemento Directo"
# 48: Estructuras gramaticales similares total
# 49: Estructuras gramaticales similares - Categoría "Preguntas con Complemento de Lugar"
# 50: Estructuras gramaticales similares - Categoría "Preguntas con Complemento de Tiempo"
# 51: Estructuras gramaticales similares - Categoría "Preguntas simples"
# 52: Estructuras gramaticales similares - Categoría "Enunciados con Complemento Directo"
# 53: Estructuras gramaticales similares distintas
# 54: Estructuras gramaticales similares distintas - Categoría "Preguntas con Complemento de Lugar"
# 55: Estructuras gramaticales similares distintas - Categoría "Preguntas con Complemento de Tiempo"
# 56: Estructuras gramaticales similares distintas - Categoría "Preguntas simples"
# 57: Estructuras gramaticales similares distintas - Categoría "Enunciados con Complemento Directo"
# 58: Actividades propuestas en pareja
# 59: Actividades acordadas en pareja
# 60: Actividades mal acordadas en pareja
''' Las siguientes estadísticas sólo están disponibles para estadísticas por alumno '''
# 61: Actividades propuestas en trío
# 62: Actividades acordadas en trío
# 63: Actividades mal acordadas en trío
# 64: Actividades acordadas individualmente


for alumno in alumnos:
    stats_per_student[alumno] = [["Mensajes", ], ["Frases", ], ["Turnos (unidades)", 0],
                                 ["Palabras", ], ["Palabras en diccionario", ], ["Palabras fuera diccionario", ],
                                 ["Palabras distintas", ], ["Palabras en diccionario distintas", ], ["Palabras fuera diccionario distintas", ],
                                 ["Palabras reservadas total", ],
                                 ["Palabras reservadas - Categoría \"Pronombres interrogativos\"", ],
                                 ["Palabras reservadas - Categoría \"Adjetivos\"", ],
                                 ["Palabras reservadas - Categoría \"Saludos y despedidas\"", ],
                                 ["Palabras reservadas - Categoría \"Frases hechas\"", ],
                                 ["Palabras reservadas - Categoría \"Expresiones de afirmación\"", ],
                                 ["Palabras reservadas - Categoría \"Expresiones de negación/duda\"", ],
                                 ["Palabras reservadas - Categoría \"Lugares\"", ],
                                 ["Palabras reservadas - Categoría \"Días de la semana\"", ],
                                 ["Palabras reservadas - Categoría \"Adverbios temporales\"", ],
                                 ["Palabras reservadas - Categoría \"Horas\"", ],
                                 ["Palabras reservadas - Categoría \"Preposición 'mit'\"", ],
                                 ["Palabras reservadas - Categoría \"'Mit' + pronombres personales\"", ],
                                 ["Palabras reservadas distintas", ],
                                 ["Palabras reservadas distintas - Categoría \"Pronombres interrogativos\"", ],
                                 ["Palabras reservadas distintas - Categoría \"Adjetivos\"", ],
                                 ["Palabras reservadas distintas - Categoría \"Saludos y despedidas\"", ],
                                 ["Palabras reservadas distintas - Categoría \"Frases hechas\"", ],
                                 ["Palabras reservadas distintas - Categoría \"Expresiones de afirmación\"", ],
                                 ["Palabras reservadas distintas - Categoría \"Expresiones de negación/duda\"", ],
                                 ["Palabras reservadas distintas - Categoría \"Lugares\"", ],
                                 ["Palabras reservadas distintas - Categoría \"Días de la semana\"", ],
                                 ["Palabras reservadas distintas - Categoría \"Adverbios temporales\"", ],
                                 ["Palabras reservadas distintas - Categoría \"Horas\"", ],
                                 ["Palabras reservadas distintas - Categoría \"Preposición 'mit'\"", ],
                                 ["Palabras reservadas distintas - Categoría \"'Mit' + pronombres personales\"", ],
                                 ["Frases de una sola palabra", ], ["Frases interrogativas", ], ["Frases exclamativas", ],
                                 ["Estructuras gramaticales precisas total", ],
                                 ["Estructuras gramaticales precisas - Categoría \"Preguntas con Complemento de Lugar\"", ],
                                 ["Estructuras gramaticales precisas - Categoría \"Preguntas con Complemento de Tiempo\"", ],
                                 ["Estructuras gramaticales precisas - Categoría \"Preguntas simples\"", ],
                                 ["Estructuras gramaticales precisas - Categoría \"Enunciados con Complemento Directo\"", ],
                                 ["Estructuras gramaticales precisas distintas", ],
                                 ["Estructuras gramaticales precisas distintas - Categoría \"Preguntas con Complemento de Lugar\"", ],
                                 ["Estructuras gramaticales precisas distintas - Categoría \"Preguntas con Complemento de Tiempo\"", ],
                                 ["Estructuras gramaticales precisas distintas - Categoría \"Preguntas simples\"", ],
                                 ["Estructuras gramaticales precisas distintas - Categoría \"Enunciados con Complemento Directo\"", ],
                                 ["Estructuras gramaticales similares total", ],
                                 ["Estructuras gramaticales similares - Categoría \"Preguntas con Complemento de Lugar\"", ],
                                 ["Estructuras gramaticales similares - Categoría \"Preguntas con Complemento de Tiempo\"", ],
                                 ["Estructuras gramaticales similares - Categoría \"Preguntas simples\"", ],
                                 ["Estructuras gramaticales similares - Categoría \"Enunciados con Complemento Directo\"", ],
                                 ["Estructuras gramaticales similares distintas", ],
                                 ["Estructuras gramaticales similares distintas - Categoría \"Preguntas con Complemento de Lugar\"", ],
                                 ["Estructuras gramaticales similares distintas - Categoría \"Preguntas con Complemento de Tiempo\"", ],
                                 ["Estructuras gramaticales similares distintas - Categoría \"Preguntas simples\"", ],
                                 ["Estructuras gramaticales similares distintas - Categoría \"Enunciados con Complemento Directo\"", ],
                                 ["Actividades propuestas en pareja", ],
                                 ["Actividades acordadas en pareja", ],
                                 ["Actividades mal acordadas en pareja", ],
                                 ["Actividades propuestas en trio", ],
                                 ["Actividades acordadas en trio", ],
                                 ["Actividades mal acordadas en trio", ],
                                 ["Actividades acordadas individualmente", ],
                                 ]

for pareja in parejas:
    stats_per_pair[pareja] = [["Mensajes", ], ["Frases", ], ["Turnos (unidades)", 0],
                              ["Palabras", ], ["Palabras en diccionario", ], ["Palabras fuera diccionario", ],
                              ["Palabras distintas", ], ["Palabras en diccionario distintas", ], ["Palabras fuera diccionario distintas", ],
                              ["Palabras reservadas total", ],
                              ["Palabras reservadas - Categoría \"Pronombres interrogativos\"", ],
                              ["Palabras reservadas - Categoría \"Adjetivos\"", ],
                              ["Palabras reservadas - Categoría \"Saludos y despedidas\"", ],
                              ["Palabras reservadas - Categoría \"Frases hechas\"", ],
                              ["Palabras reservadas - Categoría \"Expresiones de afirmación\"", ],
                              ["Palabras reservadas - Categoría \"Expresiones de negación/duda\"", ],
                              ["Palabras reservadas - Categoría \"Lugares\"", ],
                              ["Palabras reservadas - Categoría \"Días de la semana\"", ],
                              ["Palabras reservadas - Categoría \"Adverbios temporales\"", ],
                              ["Palabras reservadas - Categoría \"Horas\"", ],
                              ["Palabras reservadas - Categoría \"Preposición 'mit'\"", ],
                              ["Palabras reservadas - Categoría \"'Mit' + pronombres personales\"", ],
                              ["Palabras reservadas distintas", ],
                              ["Palabras reservadas distintas - Categoría \"Pronombres interrogativos\"", ],
                              ["Palabras reservadas distintas - Categoría \"Adjetivos\"", ],
                              ["Palabras reservadas distintas - Categoría \"Saludos y despedidas\"", ],
                              ["Palabras reservadas distintas - Categoría \"Frases hechas\"", ],
                              ["Palabras reservadas distintas - Categoría \"Expresiones de afirmación\"", ],
                              ["Palabras reservadas distintas - Categoría \"Expresiones de negación/duda\"", ],
                              ["Palabras reservadas distintas - Categoría \"Lugares\"", ],
                              ["Palabras reservadas distintas - Categoría \"Días de la semana\"", ],
                              ["Palabras reservadas distintas - Categoría \"Adverbios temporales\"", ],
                              ["Palabras reservadas distintas - Categoría \"Horas\"", ],
                              ["Palabras reservadas distintas - Categoría \"Preposición 'mit'\"", ],
                              ["Palabras reservadas distintas - Categoría \"'Mit' + pronombres personales\"", ],
                              ["Frases de una sola palabra", ], ["Frases interrogativas", ], ["Frases exclamativas", ],
                              ["Estructuras gramaticales precisas total", ],
                              ["Estructuras gramaticales precisas - Categoría \"Preguntas con Complemento de Lugar\"", ],
                              ["Estructuras gramaticales precisas - Categoría \"Preguntas con Complemento de Tiempo\"", ],
                              ["Estructuras gramaticales precisas - Categoría \"Preguntas simples\"", ],
                              ["Estructuras gramaticales precisas - Categoría \"Enunciados con Complemento Directo\"", ],
                              ["Estructuras gramaticales precisas distintas", ],
                              ["Estructuras gramaticales precisas distintas - Categoría \"Preguntas con Complemento de Lugar\"", ],
                              ["Estructuras gramaticales precisas distintas - Categoría \"Preguntas con Complemento de Tiempo\"", ],
                              ["Estructuras gramaticales precisas distintas - Categoría \"Preguntas simples\"", ],
                              ["Estructuras gramaticales precisas distintas - Categoría \"Enunciados con Complemento Directo\"", ],
                              ["Estructuras gramaticales similares total", ],
                              ["Estructuras gramaticales similares - Categoría \"Preguntas con Complemento de Lugar\"", ],
                              ["Estructuras gramaticales similares - Categoría \"Preguntas con Complemento de Tiempo\"", ],
                              ["Estructuras gramaticales similares - Categoría \"Preguntas simples\"", ],
                              ["Estructuras gramaticales similares - Categoría \"Enunciados con Complemento Directo\"", ],
                              ["Estructuras gramaticales similares distintas", ],
                              ["Estructuras gramaticales similares distintas - Categoría \"Preguntas con Complemento de Lugar\"", ],
                              ["Estructuras gramaticales similares distintas - Categoría \"Preguntas con Complemento de Tiempo\"", ],
                              ["Estructuras gramaticales similares distintas - Categoría \"Preguntas simples\"", ],
                              ["Estructuras gramaticales similares distintas - Categoría \"Enunciados con Complemento Directo\"", ],
                              ["Actividades propuestas en pareja", ],
                              ["Actividades acordadas en pareja", ],
                              ["Actividades mal acordadas en pareja", ],
                              ]

# Obtención de mensajes, frases y turnos
mensajes = []
frases = []
frases_una_palabra = []
frases_interrogativas = []
frases_exclamativas = []
numero_turnos = 0
tiempo_turnos = []

# Palabras y corrección según diccionario
palabras = []
palabras_en_diccionario = []
palabras_fuera_diccionario = []
palabras_distintas = []
palabras_en_diccionario_distintas = []
palabras_fuera_diccionario_distintas = []
caracteres_especiales = ["?", "!", ",", ".", "#", "*", "O.o", "(", ")"]
d = enchant.Dict("de_DE")

# Palabras reservadas
palabras_reservadas_total = []
palabras_reservadas_catPronombresInterrogativos = []
palabras_reservadas_catAdjetivos = []
palabras_reservadas_catSaludosDespedidas = []
palabras_reservadas_catFrasesHechas = []
palabras_reservadas_catExpresionesAfirmacion = []
palabras_reservadas_catExpresionesNegacionDuda = []
palabras_reservadas_catLugares = []
palabras_reservadas_catDiasSemana = []
palabras_reservadas_catAdverbiosTemporales = []
palabras_reservadas_catHoras = []
palabras_reservadas_catMit = []
palabras_reservadas_catPronombresPersonales = []
palabras_reservadas_total_por_categoria = [palabras_reservadas_catPronombresInterrogativos,
                                           palabras_reservadas_catAdjetivos,
                                           palabras_reservadas_catSaludosDespedidas,
                                           palabras_reservadas_catFrasesHechas,
                                           palabras_reservadas_catExpresionesAfirmacion,
                                           palabras_reservadas_catExpresionesNegacionDuda,
                                           palabras_reservadas_catLugares,
                                           palabras_reservadas_catDiasSemana,
                                           palabras_reservadas_catAdverbiosTemporales,
                                           palabras_reservadas_catHoras,
                                           palabras_reservadas_catMit,
                                           palabras_reservadas_catPronombresPersonales]

palabras_reservadas_distintas = []
palabras_reservadas_catPronombresInterrogativos_distintas = []
palabras_reservadas_catAdjetivos_distintas = []
palabras_reservadas_catSaludosDespedidas_distintas = []
palabras_reservadas_catFrasesHechas_distintas = []
palabras_reservadas_catExpresionesAfirmacion_distintas = []
palabras_reservadas_catExpresionesNegacionDuda_distintas = []
palabras_reservadas_catLugares_distintas = []
palabras_reservadas_catDiasSemana_distintas = []
palabras_reservadas_catAdverbiosTemporales_distintas = []
palabras_reservadas_catHoras_distintas = []
palabras_reservadas_catMit_distintas = []
palabras_reservadas_catPronombresPersonales_distintas = []
palabras_reservadas_distintas_por_categoria = [palabras_reservadas_catPronombresInterrogativos_distintas,
                                               palabras_reservadas_catAdjetivos_distintas,
                                               palabras_reservadas_catSaludosDespedidas_distintas,
                                               palabras_reservadas_catFrasesHechas_distintas,
                                               palabras_reservadas_catExpresionesAfirmacion_distintas,
                                               palabras_reservadas_catExpresionesNegacionDuda_distintas,
                                               palabras_reservadas_catLugares_distintas,
                                               palabras_reservadas_catDiasSemana_distintas,
                                               palabras_reservadas_catAdverbiosTemporales_distintas,
                                               palabras_reservadas_catHoras_distintas,
                                               palabras_reservadas_catMit_distintas,
                                               palabras_reservadas_catPronombresPersonales_distintas]

# Estructuras gramaticales
estructuras_gramaticales_precisas_total = []
estructuras_gramaticales_precisas_catPreguntasCLugar = []
estructuras_gramaticales_precisas_catPreguntasCTiempo = []
estructuras_gramaticales_precisas_catPreguntasSimples = []
estructuras_gramaticales_precisas_catEnunciadosCDirecto = []
estructuras_gramaticales_precisas_total_por_categoria = [estructuras_gramaticales_precisas_catPreguntasCLugar,
                                                         estructuras_gramaticales_precisas_catPreguntasCTiempo,
                                                         estructuras_gramaticales_precisas_catPreguntasSimples,
                                                         estructuras_gramaticales_precisas_catEnunciadosCDirecto]

estructuras_gramaticales_precisas_distintas = []
estructuras_gramaticales_precisas_catPreguntasCLugar_distintas = []
estructuras_gramaticales_precisas_catPreguntasCTiempo_distintas = []
estructuras_gramaticales_precisas_catPreguntasSimples_distintas = []
estructuras_gramaticales_precisas_catEnunciadosCDirecto_distintas = []
estructuras_gramaticales_precisas_distintas_por_categoria = [estructuras_gramaticales_precisas_catPreguntasCLugar_distintas,
                                                             estructuras_gramaticales_precisas_catPreguntasCTiempo_distintas,
                                                             estructuras_gramaticales_precisas_catPreguntasSimples_distintas,
                                                             estructuras_gramaticales_precisas_catEnunciadosCDirecto_distintas]

estructuras_gramaticales_similares_total = []
estructuras_gramaticales_similares_catPreguntasCLugar = []
estructuras_gramaticales_similares_catPreguntasCTiempo = []
estructuras_gramaticales_similares_catPreguntasSimples = []
estructuras_gramaticales_similares_catEnunciadosCDirecto = []
estructuras_gramaticales_similares_total_por_categoria = [estructuras_gramaticales_similares_catPreguntasCLugar,
                                                         estructuras_gramaticales_similares_catPreguntasCTiempo,
                                                         estructuras_gramaticales_similares_catPreguntasSimples,
                                                         estructuras_gramaticales_similares_catEnunciadosCDirecto]

estructuras_gramaticales_similares_distintas = []
estructuras_gramaticales_similares_catPreguntasCLugar_distintas = []
estructuras_gramaticales_similares_catPreguntasCTiempo_distintas = []
estructuras_gramaticales_similares_catPreguntasSimples_distintas = []
estructuras_gramaticales_similares_catEnunciadosCDirecto_distintas = []
estructuras_gramaticales_similares_distintas_por_categoria = [estructuras_gramaticales_similares_catPreguntasCLugar_distintas,
                                                             estructuras_gramaticales_similares_catPreguntasCTiempo_distintas,
                                                             estructuras_gramaticales_similares_catPreguntasSimples_distintas,
                                                             estructuras_gramaticales_similares_catEnunciadosCDirecto_distintas]

# Actividades en pareja
actividades_propuestas_pareja = []
actividades_acordadas_pareja = []
actividades_mal_acordadas_pareja = []

# Actividades en trío
actividades_propuestas_trio = []
actividades_acordadas_trio = []
actividades_mal_acordadas_trio = []

# Actividades individuales (solo estadísticas general y por alumno)
actividades_acordadas_individuales = []

for mensaje in Mensajes:
    # Mensajes
    mensajes.append(mensaje.mensaje)
    stats_per_student[mensaje.alumno][0].append(mensaje.mensaje)
    stats_per_pair[mensaje.pareja][0].append(mensaje.mensaje)

    # Frases
    frases_del_mensaje = []
    if "!" in mensaje.mensaje:
        if mensaje.mensaje.index("!") != len(mensaje.mensaje) - 1:
            frase = mensaje.mensaje.split("! ")
            for f in frase:
                if frase.index(f) != len(frase) - 1:
                    frases.append(f + "!")
                    stats_per_student[mensaje.alumno][1].append(f + "!")
                    stats_per_pair[mensaje.pareja][1].append(f + "!")
                    frases_del_mensaje.append(f + "!")
                else:
                    frases.append(f)
                    stats_per_student[mensaje.alumno][1].append(f)
                    stats_per_pair[mensaje.pareja][1].append(f)
                    frases_del_mensaje.append(f)
        else:
            frases.append(mensaje.mensaje)
            stats_per_student[mensaje.alumno][1].append(mensaje.mensaje)
            stats_per_pair[mensaje.pareja][1].append(mensaje.mensaje)
            frases_del_mensaje.append(mensaje.mensaje)

    if "?" in mensaje.mensaje:
        if mensaje.mensaje.index("?") != len(mensaje.mensaje) - 1:
            frase = mensaje.mensaje.split("? ")
            for f in frase:
                if frase.index(f) != len(frase) - 1:
                    frases.append(f + "?")
                    stats_per_student[mensaje.alumno][1].append(f + "?")
                    stats_per_pair[mensaje.pareja][1].append(f + "?")
                    frases_del_mensaje.append(f + "?")
                else:
                    frases.append(f)
                    stats_per_student[mensaje.alumno][1].append(f)
                    stats_per_pair[mensaje.pareja][1].append(f)
                    frases_del_mensaje.append(f)
        elif "!" not in mensaje.mensaje:
            frases.append(mensaje.mensaje)
            stats_per_student[mensaje.alumno][1].append(mensaje.mensaje)
            stats_per_pair[mensaje.pareja][1].append(mensaje.mensaje)
            frases_del_mensaje.append(mensaje.mensaje)

    if "." in mensaje.mensaje and "..." not in mensaje.mensaje:
        if mensaje.mensaje.index(".") != len(mensaje.mensaje) - 1:
            frase = mensaje.mensaje.split(". ")
            for f in frase:
                if frase.index(f) != len(frase) - 1:
                    frases.append(f + ".")
                    stats_per_student[mensaje.alumno][1].append(f + ".")
                    stats_per_pair[mensaje.pareja][1].append(f + ".")
                    frases_del_mensaje.append(f + ".")
                else:
                    frases.append(f)
                    stats_per_student[mensaje.alumno][1].append(f)
                    stats_per_pair[mensaje.pareja][1].append(f)
                    frases_del_mensaje.append(f)
        elif "!" not in mensaje.mensaje and "?" not in mensaje.mensaje:
            frases.append(mensaje.mensaje)
            stats_per_student[mensaje.alumno][1].append(mensaje.mensaje)
            stats_per_pair[mensaje.pareja][1].append(mensaje.mensaje)
            frases_del_mensaje.append(mensaje.mensaje)

    if "!" not in mensaje.mensaje and "?" not in mensaje.mensaje and "." not in mensaje.mensaje:
        frases.append(mensaje.mensaje)
        stats_per_student[mensaje.alumno][1].append(mensaje.mensaje)
        stats_per_pair[mensaje.pareja][1].append(mensaje.mensaje)
        frases_del_mensaje.append(mensaje.mensaje)

    for frase in frases_del_mensaje:
        if len(frase.split()) == 1 or (len(frase.split()) == 2 and ('?' in frase or '!' in frase)):
            frases_una_palabra.append(frase)
            stats_per_student[mensaje.alumno][35].append(frase)
            stats_per_pair[mensaje.pareja][35].append(frase)

        if '?' in frase:
            frases_interrogativas.append(frase)
            stats_per_student[mensaje.alumno][36].append(frase)
            stats_per_pair[mensaje.pareja][36].append(frase)

        if '!' in frase:
            frases_exclamativas.append(frase)
            stats_per_student[mensaje.alumno][37].append(frase)
            stats_per_pair[mensaje.pareja][37].append(frase)

    # Turnos
    if mensaje.cambio_turno:
        numero_turnos += 1
        stats_per_student[mensaje.alumno][2][1] += 1
        stats_per_pair[mensaje.pareja][2][1] += 1

    # Diccionario
    lista_palabras = mensaje.mensaje.split()
    for palabra in lista_palabras:
        if palabra not in caracteres_especiales:
            # Palabras
            palabras.append(palabra)
            if palabra not in palabras_distintas:
                palabras_distintas.append(palabra)

            stats_per_student[mensaje.alumno][3].append(palabra)
            if palabra not in stats_per_student[mensaje.alumno][6]:
                stats_per_student[mensaje.alumno][6].append(palabra)

            stats_per_pair[mensaje.pareja][3].append(palabra)
            if palabra not in stats_per_pair[mensaje.pareja][6]:
                stats_per_pair[mensaje.pareja][6].append(palabra)

            # Palabras correctas/incorrectas
            if d.check(palabra) or palabra.lower() in ampliacion_diccionario or palabra.capitalize() in ampliacion_diccionario:
                palabras_en_diccionario.append(palabra)
                if palabra not in palabras_en_diccionario_distintas:
                    palabras_en_diccionario_distintas.append(palabra)

                stats_per_student[mensaje.alumno][4].append(palabra)
                if palabra not in stats_per_student[mensaje.alumno][7]:
                    stats_per_student[mensaje.alumno][7].append(palabra)

                stats_per_pair[mensaje.pareja][4].append(palabra)
                if palabra not in stats_per_pair[mensaje.pareja][7]:
                    stats_per_pair[mensaje.pareja][7].append(palabra)
            else:
                palabras_fuera_diccionario.append(palabra)
                if palabra not in palabras_fuera_diccionario_distintas:
                    palabras_fuera_diccionario_distintas.append(palabra)

                stats_per_student[mensaje.alumno][5].append(palabra)
                if palabra not in stats_per_student[mensaje.alumno][8]:
                    stats_per_student[mensaje.alumno][8].append(palabra)

                stats_per_pair[mensaje.pareja][5].append(palabra)
                if palabra not in stats_per_pair[mensaje.pareja][8]:
                    stats_per_pair[mensaje.pareja][8].append(palabra)

    # Palabras reservadas
    expresiones_en_mensaje = usaExpresionReservada(mensaje.mensaje)
    if len(expresiones_en_mensaje) >= 1:
        for expresion in expresiones_en_mensaje:
            palabras_reservadas_total.append(expresion)
            if expresion not in palabras_reservadas_distintas:
                palabras_reservadas_distintas.append(expresion)

            stats_per_student[mensaje.alumno][9].append(expresion)
            if expresion not in stats_per_student[mensaje.alumno][22]:
                stats_per_student[mensaje.alumno][22].append(expresion)

            stats_per_pair[mensaje.pareja][9].append(expresion)
            if expresion not in stats_per_pair[mensaje.pareja][22]:
                stats_per_pair[mensaje.pareja][22].append(expresion)

            tipo_expresion = expresion[0]

            if tipo_expresion == "Pronombre interrogativo":
                palabras_reservadas_catPronombresInterrogativos.append(expresion)
                if expresion not in palabras_reservadas_catPronombresInterrogativos_distintas:
                    palabras_reservadas_catPronombresInterrogativos_distintas.append(expresion)

                stats_per_student[mensaje.alumno][10].append(expresion)
                if expresion not in stats_per_student[mensaje.alumno][23]:
                    stats_per_student[mensaje.alumno][23].append(expresion)

                stats_per_pair[mensaje.pareja][10].append(expresion)
                if expresion not in stats_per_pair[mensaje.pareja][23]:
                    stats_per_pair[mensaje.pareja][23].append(expresion)

            elif tipo_expresion == "Adjetivo":
                palabras_reservadas_catAdjetivos.append(expresion)
                if expresion not in palabras_reservadas_catAdjetivos_distintas:
                    palabras_reservadas_catAdjetivos_distintas.append(expresion)

                stats_per_student[mensaje.alumno][11].append(expresion)
                if expresion not in stats_per_student[mensaje.alumno][24]:
                    stats_per_student[mensaje.alumno][24].append(expresion)

                stats_per_pair[mensaje.pareja][9].append(expresion)
                if expresion not in stats_per_pair[mensaje.pareja][22]:
                    stats_per_pair[mensaje.pareja][22].append(expresion)

            elif tipo_expresion == "Saludo o despedida":
                palabras_reservadas_catSaludosDespedidas.append(expresion)
                if expresion not in palabras_reservadas_catSaludosDespedidas_distintas:
                    palabras_reservadas_catSaludosDespedidas_distintas.append(expresion)

                stats_per_student[mensaje.alumno][12].append(expresion)
                if expresion not in stats_per_student[mensaje.alumno][25]:
                    stats_per_student[mensaje.alumno][25].append(expresion)

                stats_per_pair[mensaje.pareja][12].append(expresion)
                if expresion not in stats_per_pair[mensaje.pareja][25]:
                    stats_per_pair[mensaje.pareja][25].append(expresion)

            elif tipo_expresion == "Frase hecha":
                palabras_reservadas_catFrasesHechas.append(expresion)
                if expresion not in palabras_reservadas_catFrasesHechas_distintas:
                    palabras_reservadas_catFrasesHechas_distintas.append(expresion)

                stats_per_student[mensaje.alumno][13].append(expresion)
                if expresion not in stats_per_student[mensaje.alumno][26]:
                    stats_per_student[mensaje.alumno][26].append(expresion)

                stats_per_pair[mensaje.pareja][13].append(expresion)
                if expresion not in stats_per_pair[mensaje.pareja][26]:
                    stats_per_pair[mensaje.pareja][26].append(expresion)

            elif tipo_expresion == "Expresion de afirmacion":
                palabras_reservadas_catExpresionesAfirmacion.append(expresion)
                if expresion not in palabras_reservadas_catExpresionesAfirmacion_distintas:
                    palabras_reservadas_catExpresionesAfirmacion_distintas.append(expresion)

                stats_per_student[mensaje.alumno][14].append(expresion)
                if expresion not in stats_per_student[mensaje.alumno][27]:
                    stats_per_student[mensaje.alumno][27].append(expresion)

                stats_per_pair[mensaje.pareja][14].append(expresion)
                if expresion not in stats_per_pair[mensaje.pareja][27]:
                    stats_per_pair[mensaje.pareja][27].append(expresion)

            elif tipo_expresion == "Expresion de negacion o duda":
                palabras_reservadas_catExpresionesNegacionDuda.append(expresion)
                if expresion not in palabras_reservadas_catExpresionesNegacionDuda_distintas:
                    palabras_reservadas_catExpresionesNegacionDuda_distintas.append(expresion)

                stats_per_student[mensaje.alumno][15].append(expresion)
                if expresion not in stats_per_student[mensaje.alumno][28]:
                    stats_per_student[mensaje.alumno][28].append(expresion)

                stats_per_pair[mensaje.pareja][15].append(expresion)
                if expresion not in stats_per_pair[mensaje.pareja][28]:
                    stats_per_pair[mensaje.pareja][28].append(expresion)

            elif tipo_expresion == "Lugar":
                palabras_reservadas_catLugares.append(expresion)
                if expresion not in palabras_reservadas_catLugares_distintas:
                    palabras_reservadas_catLugares_distintas.append(expresion)

                stats_per_student[mensaje.alumno][16].append(expresion)
                if expresion not in stats_per_student[mensaje.alumno][29]:
                    stats_per_student[mensaje.alumno][29].append(expresion)

                stats_per_pair[mensaje.pareja][16].append(expresion)
                if expresion not in stats_per_pair[mensaje.pareja][29]:
                    stats_per_pair[mensaje.pareja][29].append(expresion)

            elif tipo_expresion == "Dia de la semana":
                palabras_reservadas_catDiasSemana.append(expresion)
                if expresion not in palabras_reservadas_catDiasSemana_distintas:
                    palabras_reservadas_catDiasSemana_distintas.append(expresion)

                stats_per_student[mensaje.alumno][17].append(expresion)
                if expresion not in stats_per_student[mensaje.alumno][30]:
                    stats_per_student[mensaje.alumno][30].append(expresion)

                stats_per_pair[mensaje.pareja][17].append(expresion)
                if expresion not in stats_per_pair[mensaje.pareja][30]:
                    stats_per_pair[mensaje.pareja][30].append(expresion)

            elif tipo_expresion == "Adverbio temporal":
                palabras_reservadas_catAdverbiosTemporales.append(expresion)
                if expresion not in palabras_reservadas_catAdverbiosTemporales_distintas:
                    palabras_reservadas_catAdverbiosTemporales_distintas.append(expresion)

                stats_per_student[mensaje.alumno][18].append(expresion)
                if expresion not in stats_per_student[mensaje.alumno][31]:
                    stats_per_student[mensaje.alumno][31].append(expresion)

                stats_per_pair[mensaje.pareja][18].append(expresion)
                if expresion not in stats_per_pair[mensaje.pareja][31]:
                    stats_per_pair[mensaje.pareja][31].append(expresion)

            elif tipo_expresion == "Hora":
                palabras_reservadas_catHoras.append(expresion)
                if expresion not in palabras_reservadas_catHoras_distintas:
                    palabras_reservadas_catHoras_distintas.append(expresion)

                stats_per_student[mensaje.alumno][19].append(expresion)
                if expresion not in stats_per_student[mensaje.alumno][32]:
                    stats_per_student[mensaje.alumno][32].append(expresion)

                stats_per_pair[mensaje.pareja][19].append(expresion)
                if expresion not in stats_per_pair[mensaje.pareja][32]:
                    stats_per_pair[mensaje.pareja][32].append(expresion)

            elif tipo_expresion == "Preposicion mit":
                palabras_reservadas_catMit.append(expresion)
                if expresion not in palabras_reservadas_catMit_distintas:
                    palabras_reservadas_catMit_distintas.append(expresion)

                stats_per_student[mensaje.alumno][20].append(expresion)
                if expresion not in stats_per_student[mensaje.alumno][33]:
                    stats_per_student[mensaje.alumno][33].append(expresion)

                stats_per_pair[mensaje.pareja][20].append(expresion)
                if expresion not in stats_per_pair[mensaje.pareja][33]:
                    stats_per_pair[mensaje.pareja][33].append(expresion)

            elif tipo_expresion == "Preposicion con pronombre personal":
                palabras_reservadas_catPronombresPersonales.append(expresion)
                if expresion not in palabras_reservadas_catPronombresPersonales_distintas:
                    palabras_reservadas_catPronombresPersonales_distintas.append(expresion)

                stats_per_student[mensaje.alumno][21].append(expresion)
                if expresion not in stats_per_student[mensaje.alumno][34]:
                    stats_per_student[mensaje.alumno][34].append(expresion)

                stats_per_pair[mensaje.pareja][21].append(expresion)
                if expresion not in stats_per_pair[mensaje.pareja][34]:
                    stats_per_pair[mensaje.pareja][34].append(expresion)

            else:
                print "Esto es más raro que la historia de Manuel Bartual..."


# Extracción de actividades
agendas = []

for fichero in os.listdir("."):
    if "Agenda" in fichero and "#" not in fichero:
        agendas.append(fichero)

for ag in agendas:
    agenda = load_workbook(ag)

# Salida a fichero de texto
file = open("stats.txt", "w")

file.write("__________________________________\n")
file.write("|                                |\n")
file.write("| ESTADÍSTICAS EXTRAÍDAS (PAPER) |\n")
file.write("|________________________________|\n")

file.write("\n-------> Categoría: General <-------\n")
file.write("\nPalabras: " + str(len(palabras)))
file.write("\nPalabras distintas: " + str(len(palabras_distintas)))
file.write("\nPalabras correctas: " + str(len(palabras_en_diccionario)))
file.write("\nPalabras correctas distintas: " + str(len(palabras_en_diccionario_distintas)))
file.write("\nPalabras incorrectas: " + str(len(palabras_fuera_diccionario)))
file.write("\nPalabras incorrectas distintas: " + str(len(palabras_fuera_diccionario_distintas)))
file.write("\nPalabras reservadas total: " + str(len(palabras_reservadas_total)))
file.write("\nPalabras reservadas distintas: " + str(len(palabras_reservadas_distintas)))
i = 0
while i < len(palabras_reservadas_total_por_categoria):
    file.write("\nPalabras reservadas total - Categoría " + palabras_reservadas_total_por_categoria[i][0][0] + ": " + str(len(palabras_reservadas_total_por_categoria[i])))
    file.write("\n... de las cuales distintas: " + str(len(palabras_reservadas_distintas_por_categoria[i])))
    i += 1
file.write("\nFrases: " + str(len(frases)))
file.write("\nFrases de una sola palabra: " + str(len(frases_una_palabra)))
file.write("\nFrases interrogativas: " + str(len(frases_interrogativas)))
file.write("\nFrases exclamativas: " + str(len(frases_exclamativas)))
file.write("\nMensajes: " + str(len(mensajes)))
file.write("\nTurnos (unidades): " + str(numero_turnos))
'''file.write("\n\n-- Actividades --")
file.write("\nActividades individuales:")
file.write("\n-> Acordadas: " + str(len(actividades_individuales)))
file.write("\nActividades en pareja:")
file.write("\n-> Propuestas: " + str(len(actividades_propuestas_pareja)))
file.write("\n-> Acordadas: " + str(len(actividades_acordadas_pareja)))
file.write("\n-> Mal acordadas: " + str(len(actividades_mal_acordadas_pareja)))
file.write("\nActividades en trío:")
file.write("\n-> Propuestas: " + str(len(actividades_propuestas_trio)))
file.write("\n-> Acordadas: " + str(len(actividades_acordadas_trio)))
file.write("\n-> Mal acordadas: " + str(len(actividades_mal_acordadas_trio)))'''

file.write("\n\n-------> Categoría: Por alumno <-------\n")
for alumno in alumnos:
    file.write("\n\n==> " + alumno)
    file.write("\nPalabras: " + str(len(stats_per_student[alumno][3]) - 1))
    file.write("\nPalabras distintas: " + str(len(stats_per_student[alumno][6]) - 1))
    file.write("\nPalabras correctas: " + str(len(stats_per_student[alumno][4]) - 1))
    file.write("\nPalabras correctas distintas: " + str(len(stats_per_student[alumno][7]) - 1))
    file.write("\nPalabras incorrectas: " + str(len(stats_per_student[alumno][5]) - 1))
    file.write("\nPalabras incorrectas distintas: " + str(len(stats_per_student[alumno][8]) - 1))
    file.write("\nPalabras reservadas total: " + str(len(stats_per_student[alumno][9]) - 1))
    file.write("\nPalabras reservadas distintas: " + str(len(stats_per_student[alumno][22]) - 1))
    file.write("\nPalabras reservadas total - Categoría Pronombre Interrogativo: " + str(len(stats_per_student[alumno][10]) - 1))
    file.write("\n... de las cuales distintas: " + str(len(stats_per_student[alumno][23]) - 1))
    file.write("\nPalabras reservadas total - Categoría Adjetivo: " + str(len(stats_per_student[alumno][11]) - 1))
    file.write("\n... de las cuales distintas: " + str(len(stats_per_student[alumno][24]) - 1))
    file.write("\nPalabras reservadas total - Categoría Saludo o Despedida: " + str(len(stats_per_student[alumno][12]) - 1))
    file.write("\n... de las cuales distintas: " + str(len(stats_per_student[alumno][25]) - 1))
    file.write("\nPalabras reservadas total - Categoría Frase hecha: " + str(len(stats_per_student[alumno][13]) - 1))
    file.write("\n... de las cuales distintas: " + str(len(stats_per_student[alumno][26]) - 1))
    file.write("\nPalabras reservadas total - Categoría Expresión de afirmación: " + str(len(stats_per_student[alumno][14]) - 1))
    file.write("\n... de las cuales distintas: " + str(len(stats_per_student[alumno][27]) - 1))
    file.write("\nPalabras reservadas total - Categoría Expresión de negación/duda: " + str(len(stats_per_student[alumno][15]) - 1))
    file.write("\n... de las cuales distintas: " + str(len(stats_per_student[alumno][28]) - 1))
    file.write("\nPalabras reservadas total - Categoría Lugar: " + str(len(stats_per_student[alumno][16]) - 1))
    file.write("\n... de las cuales distintas: " + str(len(stats_per_student[alumno][29]) - 1))
    file.write("\nPalabras reservadas total - Categoría Día de la semana: " + str(len(stats_per_student[alumno][17]) - 1))
    file.write("\n... de las cuales distintas: " + str(len(stats_per_student[alumno][30]) - 1))
    file.write("\nPalabras reservadas total - Categoría Adverbio temporal: " + str(len(stats_per_student[alumno][18]) - 1))
    file.write("\n... de las cuales distintas: " + str(len(stats_per_student[alumno][31]) - 1))
    file.write("\nPalabras reservadas total - Categoría Hora: " + str(len(stats_per_student[alumno][19]) - 1))
    file.write("\n... de las cuales distintas: " + str(len(stats_per_student[alumno][32]) - 1))
    file.write("\nPalabras reservadas total - Categoría Preposición \'mit\': " + str(len(stats_per_student[alumno][20]) - 1))
    file.write("\n... de las cuales distintas: " + str(len(stats_per_student[alumno][33]) - 1))
    file.write("\nPalabras reservadas total - Categoría \'Mit\' + Pronombre personal: " + str(len(stats_per_student[alumno][21]) - 1))
    file.write("\n... de las cuales distintas: " + str(len(stats_per_student[alumno][34]) - 1))
    file.write("\nFrases: " + str(len(stats_per_student[alumno][1]) - 1))
    file.write("\nFrases de una sola palabra: " + str(len(stats_per_student[alumno][35]) - 1))
    file.write("\nFrases interrogativas: " + str(len(stats_per_student[alumno][36]) - 1))
    file.write("\nFrases exclamativas: " + str(len(stats_per_student[alumno][37]) - 1))
    file.write("\nMensajes: " + str(len(stats_per_student[alumno][0]) - 1))
    file.write("\nTurnos (unidades): " + str(stats_per_student[alumno][2][1]))
    '''file.write("\n\n-- Actividades --")
    file.write("\nActividades individuales:")
    file.write("\n-> Acordadas: " + str(len(stats_per_student[alumno][64]) - 1))
    file.write("\nActividades en pareja:")
    file.write("\n-> Pendientes de respuesta: " + str(len(stats_per_student[alumno][58]) - 1))
    file.write("\n-> Acordadas: " + str(len(stats_per_student[alumno][59]) - 1))
    file.write("\n-> Mal acordadas: " + str(len(stats_per_student[alumno][60]) - 1))
    file.write("\nActividades en trío:")
    file.write("\n-> Pendientes de respuesta: " + str(len(stats_per_student[alumno][61]) - 1))
    file.write("\n-> Acordadas: " + str(len(stats_per_student[alumno][62]) - 1))
    file.write("\n-> Mal acordadas: " + str(len(stats_per_student[alumno][63]) - 1))'''

file.write("\n\n-------> Categoría: Por pareja <-------\n")
for pareja in parejas:
    file.write("\n\n==> " + pareja)
    file.write("\nPalabras: " + str(len(stats_per_pair[pareja][3]) - 1))
    file.write("\nPalabras distintas: " + str(len(stats_per_pair[pareja][6]) - 1))
    file.write("\nPalabras correctas: " + str(len(stats_per_pair[pareja][4]) - 1))
    file.write("\nPalabras correctas distintas: " + str(len(stats_per_pair[pareja][7]) - 1))
    file.write("\nPalabras incorrectas: " + str(len(stats_per_pair[pareja][5]) - 1))
    file.write("\nPalabras incorrectas distintas: " + str(len(stats_per_pair[pareja][8]) - 1))
    file.write("\nPalabras reservadas total: " + str(len(stats_per_pair[pareja][9]) - 1))
    file.write("\nPalabras reservadas distintas: " + str(len(stats_per_pair[pareja][22]) - 1))
    file.write("\nPalabras reservadas total - Categoría Pronombre Interrogativo: " + str(len(stats_per_pair[pareja][10]) - 1))
    file.write("\n... de las cuales distintas: " + str(len(stats_per_pair[pareja][23]) - 1))
    file.write("\nPalabras reservadas total - Categoría Adjetivo: " + str(len(stats_per_pair[pareja][11]) - 1))
    file.write("\n... de las cuales distintas: " + str(len(stats_per_pair[pareja][24]) - 1))
    file.write("\nPalabras reservadas total - Categoría Saludo o Despedida: " + str(len(stats_per_pair[pareja][12]) - 1))
    file.write("\n... de las cuales distintas: " + str(len(stats_per_pair[pareja][25]) - 1))
    file.write("\nPalabras reservadas total - Categoría Frase hecha: " + str(len(stats_per_pair[pareja][13]) - 1))
    file.write("\n... de las cuales distintas: " + str(len(stats_per_pair[pareja][26]) - 1))
    file.write("\nPalabras reservadas total - Categoría Expresión de afirmación: " + str(len(stats_per_pair[pareja][14]) - 1))
    file.write("\n... de las cuales distintas: " + str(len(stats_per_pair[pareja][27]) - 1))
    file.write("\nPalabras reservadas total - Categoría Expresión de negación/duda: " + str(len(stats_per_pair[pareja][15]) - 1))
    file.write("\n... de las cuales distintas: " + str(len(stats_per_pair[pareja][28]) - 1))
    file.write("\nPalabras reservadas total - Categoría Lugar: " + str(len(stats_per_pair[pareja][16]) - 1))
    file.write("\n... de las cuales distintas: " + str(len(stats_per_pair[pareja][29]) - 1))
    file.write("\nPalabras reservadas total - Categoría Día de la semana: " + str(len(stats_per_pair[pareja][17]) - 1))
    file.write("\n... de las cuales distintas: " + str(len(stats_per_pair[pareja][30]) - 1))
    file.write("\nPalabras reservadas total - Categoría Adverbio temporal: " + str(len(stats_per_pair[pareja][18]) - 1))
    file.write("\n... de las cuales distintas: " + str(len(stats_per_pair[pareja][31]) - 1))
    file.write("\nPalabras reservadas total - Categoría Hora: " + str(len(stats_per_pair[pareja][19]) - 1))
    file.write("\n... de las cuales distintas: " + str(len(stats_per_pair[pareja][32]) - 1))
    file.write("\nPalabras reservadas total - Categoría Preposición \'mit\': " + str(len(stats_per_pair[pareja][20]) - 1))
    file.write("\n... de las cuales distintas: " + str(len(stats_per_pair[pareja][33]) - 1))
    file.write("\nPalabras reservadas total - Categoría \'Mit\' + Pronombre personal: " + str(len(stats_per_pair[pareja][21]) - 1))
    file.write("\n... de las cuales distintas: " + str(len(stats_per_pair[pareja][34]) - 1))
    file.write("\nFrases: " + str(len(stats_per_pair[pareja][1]) - 1))
    file.write("\nFrases de una sola palabra: " + str(len(stats_per_pair[pareja][35]) - 1))
    file.write("\nFrases interrogativas: " + str(len(stats_per_pair[pareja][36]) - 1))
    file.write("\nFrases exclamativas: " + str(len(stats_per_pair[pareja][37]) - 1))
    file.write("\nMensajes: " + str(len(stats_per_pair[pareja][0]) - 1))
    file.write("\nTurnos (unidades): " + str(stats_per_pair[pareja][2][1]))
    file.write("\n\n-- Actividades --")
    '''file.write("\nActividades en pareja:")
    file.write("\n-> Pendientes de respuesta: " + str(len(stats_per_pair[pareja][58]) - 1))
    file.write("\n-> Acordadas: " + str(len(stats_per_pair[pareja][59]) - 1))
    file.write("\n-> Mal acordadas: " + str(len(stats_per_pair[pareja][60]) - 1))

file.write("\n\n-------> Categoría: Por trío <-------\n")
for trio in trios:
    file.write("\n\n==> " + trio)
    file.write("\n\n-- Actividades --")
    file.write("\nActividades en trío:")
    file.write("\n-> Pendientes de respuesta: " + str(len(actividades_por_trios[trio][0]) - 1))
    file.write("\n-> Acordadas: " + str(len(actividades_por_trios[trio][1]) - 1))
    file.write("\n-> Mal acordadas: " + str(len(actividades_por_trios[trio][2]) - 1))'''

file.write("\n\n\nGracias por usar TerminKalender Stats.")
file.write("\nGithub: https://github.com/luisrozo/tk-stats\n")

print "Fichero generado con estadísticas . . ."
print "Gracias por usar TerminKalender Stats."
print "Github: https://github.com/luisrozo/tk-stats"
